import openpyxl
import os
import time
import tkinter.ttk
import tkinter.filedialog
import tkinter.messagebox
import webbrowser

root = tkinter.Tk()
root.title("업데이트 체크")
root.overrideredirect(True)
root.geometry("+%d+0" % (root.winfo_screenwidth() - 283))

filename = "settings.xlsx"  # 엑셀파일명 등록
excel = openpyxl.load_workbook(filename)  # 엑셀파일 지정
game = excel.worksheets[0]  # 시트 지정
patch = excel.worksheets[1]  # 시트 지정

name = 2  # 게임이름 열 번호
pwd = 3  # 파일 경로 열 번호
day = 4  # 날짜 열 번호


def close():
    root.quit()
    root.destroy()


menubar = tkinter.Menu(root)
menu = tkinter.Menu(menubar, tearoff=0)
menu.add_command(label="Exit", command=close)
menubar.add_cascade(label="Option", menu=menu)
root.config(menu=menubar)

treeview = tkinter.ttk.Treeview(root, columns=["one"], displaycolumns=["one"], height=game.max_row - 2)
treeview.pack()
treeview.column("#0", width=180)
treeview.heading("#0", text="게임명")
treeview.column("#1", width=100)
treeview.heading("#1", text="패치날짜")


def load():
    for column in range(3, game.max_row + 1):
        if game.cell(column, pwd).value is None:
            continue
        filestat = os.stat(game.cell(column, pwd).value)
        time_v = time.strftime('%Y/%m/%d', time.localtime(filestat.st_mtime))
        treeview.item(str(column), text=game.cell(column, name).value, values=time_v)
        game.cell(column, day).value = time_v  # 수정날짜 엑셀에 기록
    excel.save(filename='게임 패치 목록.xlsx')  # 저장


def OnDoubleClick(event):
    num = int(treeview.selection()[0])
    filename = tkinter.filedialog.askdirectory()
    if filename:
        game.cell(num, pwd).value = filename
        excel.save(filename='settings.xlsx')


def update():
    num = int(treeview.selection()[0])
    if patch.cell(num, 5).value == "Yes":
        webbrowser.open(patch.cell(num, 7).value)
    else:
        tkinter.messagebox.showinfo("메시지 상자", "넥슨 홈페이지에서 시도하세요!")


for i in range(3, game.max_row + 1):
    treeview.insert('', 'end', text=game.cell(i, name).value, values='', iid=str(i))
load()

Reload = tkinter.Button(root, overrelief="solid", command=load, repeatdelay=1000, repeatinterval=100, text="Reload")
Update = tkinter.Button(root, overrelief="solid", command=update, repeatdelay=1000, repeatinterval=100, text="Update")
treeview.bind("<Double-1>", OnDoubleClick)
Reload.pack(side="left", expand="yes", fill="x")
Update.pack(side="right", expand="yes", fill="x")
root.mainloop()
